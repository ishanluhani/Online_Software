import os
from tkinter import filedialog
from tkinter import *
from openpyxl import load_workbook
from os import listdir, path, startfile

def browse():
    global path_target
    path_target = filedialog.askdirectory()
    browse_text['text'] = path_target


def submit(column_idx, target_folder, lst):
    global l
    l.delete(0, END)
    print(lst, 'djhgjdffj')
    for x in listdir(target_folder):
        print(path.join(target_folder, x))
        try:
            file = load_workbook(path.join(target_folder, x))
            file_sheet = file.worksheets[0]
            for i in range(file_sheet.max_row):
                c = file_sheet.cell(i + 1, int(column_idx)).value
                if lst in str(c):
                    l.insert(END, f'{x}')
                    startfile(path.join(target_folder, x))
        except:
            print('error')


root = Tk()
path_target = 'D:\meesho customer database'

text_in_place_of_sub_order = 'Text To Search'.upper()
text_in_place_of_column = 'Column Number'.upper()
text_in_place_of_target = 'Select target folder'.upper()
root.geometry('400x550')
Label(root, text=text_in_place_of_sub_order).pack()
sub_entry = Entry(root, width=50)
sub_entry.pack()
Label(root, text=text_in_place_of_column).pack()
col_entry = Entry(root, width=50)
col_entry.insert(0, '5')
col_entry.pack()
Label(root, text=text_in_place_of_target).pack()
Button(root, text='Browse',command=browse).pack()
browse_text = Label(root, text=path_target)
browse_text.pack()
print(sub_entry.get(), 'fg')
Button(root, text='Submit',command=lambda: submit(col_entry.get(), path_target, sub_entry.get()), bg='green', fg='black').pack()
root.bind('<Return>', lambda x: submit(col_entry.get(), path_target, sub_entry.get()))
l = Listbox(root, width=50, height=20)
l.pack()

root.mainloop()
