from openpyxl import load_workbook
from os import path,listdir
import colorama
manual_data = []
manual_file = r'C:\neetu onedrive\OneDrive\DAILY PHYSICAL RETURNS DATA\manual'
for i in listdir(manual_file):
    if '~$' in i:
        continue
    print(path.join(manual_file, i))
    manual_workbook = load_workbook(open(path.join(manual_file, i), 'rb'))
    manual_sheet = manual_workbook.worksheets[0]
    x = 1
    while True:
        x += 1
        temp_val = manual_sheet[f'A{x}']
        temp_val1 = manual_sheet[f'D{x}']
        temp_val2 = manual_sheet[f'G{x}']
        temp_val3 = manual_sheet[f'J{x}']
        temp_val4 = manual_sheet[f'M{x}']
        temp_val5 = manual_sheet[f'P{x}']
        temp_val6 = manual_sheet[f'S{x}']
        if temp_val.value:
            manual_data.append(str(temp_val.value))
        if temp_val1.value:
            manual_data.append(str(temp_val1.value))
        if temp_val2.value:
            manual_data.append(str(temp_val2.value))
        if temp_val3.value:
            manual_data.append(str(temp_val3.value))
        if temp_val4.value:
            manual_data.append(str(temp_val4.value))
        if temp_val5.value:
            manual_data.append(str(temp_val5.value))
        if temp_val6.value:
            manual_data.append(str(temp_val6.value))
        if not any([temp_val3.value, temp_val2.value, temp_val1.value, temp_val.value, temp_val4.value, temp_val5.value, temp_val6.value]):
            break

meesho_data = []
meesho_file = r'C:\neetu onedrive\OneDrive\DAILY PHYSICAL RETURNS DATA\meesho'
temp = manual_data[:]
for meehso_p in listdir(meesho_file):
    meesho_workbook = load_workbook(path.join(meesho_file, meehso_p))
    meesho_sheet = meesho_workbook.worksheets[0]
    x = 8
    while True:
        x += 1
        an = str(meesho_sheet.cell(x,17).value)
        an2 = str(meesho_sheet.cell(x,9).value)
        if an == 'None':
            break
        meesho_data.append(an)
        meesho_data.append(an2)
        # temp = list(map(lambda x: x[0], manual_data[:]))
        if an in temp or an2 in temp:
            meesho_sheet[f'U{x}'].value = 'ok'
            print(meesho_sheet.cell(x,21).value)
    meesho_workbook.save(f'{meesho_file}/{meehso_p}')
def add_and_check(val, index, val1):
    if val.value:
        if str(val.value) in meesho_data:
            manual_sheet.cell(index, val1).value = 'ok'
        else:
            manual_sheet.cell(index, val1).value = ''
for i in listdir(manual_file):
    if '~$' in i:
        continue
    manual_workbook = load_workbook(path.join(manual_file, i))
    manual_sheet = manual_workbook.worksheets[0]
    x = 1
    while True:
        x += 1
        temp_val = manual_sheet[f'A{x}']
        temp_val1 = manual_sheet[f'D{x}']
        temp_val2 = manual_sheet[f'G{x}']
        temp_val3 = manual_sheet[f'J{x}']
        temp_val4 = manual_sheet[f'M{x}']
        temp_val5 = manual_sheet[f'P{x}']
        temp_val6 = manual_sheet[f'S{x}']
        add_and_check(temp_val, x, 3)
        add_and_check(temp_val1, x, 6)
        add_and_check(temp_val2, x, 9)
        add_and_check(temp_val3, x, 12)
        add_and_check(temp_val4, x, 15)
        add_and_check(temp_val5, x, 18)
        add_and_check(temp_val6, x, 21)
        if not any([temp_val3.value, temp_val2.value, temp_val1.value, temp_val.value, temp_val4.value, temp_val5.value, temp_val6.value]):
            break
    manual_workbook.save(f'{manual_file}/{i}')


print(colorama.Fore.GREEN, 'Process has finished please continue.')