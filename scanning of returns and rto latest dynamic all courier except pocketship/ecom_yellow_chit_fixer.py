from PyPDF2 import PdfFileReader
from tkinter.filedialog import askopenfilename
import openpyxl
from glob import glob
import os


def process(a2):
    excel = openpyxl.load_workbook(a2)
    excel_sheet = excel.worksheets[0]
    a1 = PdfFileReader(askopenfilename())

    for i in range(excel_sheet.max_row-1):
        val = excel_sheet.cell(i+2, 1).value
        print(val)
        try:
            if val[0] == '9':
                for x in range(a1.getNumPages()):
                    a = a1.getPage(x).extractText().split()
                    try:
                        excel_sheet.cell(i+2, 1).value = a[a.index(val)-3]
                        break
                    except:
                        continue
        except:
            break
    excel.save(a2)

