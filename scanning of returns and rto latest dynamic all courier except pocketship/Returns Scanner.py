import os

import pandas as pd
from openpyxl import Workbook
from tkinter import *
from tkinter.messagebox import askyesno
import pygame.mixer
import datetime
import ecom_yellow_chit_fixer

pygame.mixer.init()
pass_sound = pygame.mixer.Sound('beep-06.mp3')
error_sound = pygame.mixer.Sound('beep-03.mp3')


def update():
    listbox.delete(0, END)
    for i in scanned_data:
        listbox.insert(END, f'{i}:   {len(scanned_data[i])}')


def finish():
    n = 1
    for i in scanned_data:
        try:
            output_sheet.cell(1, n).value = i
            for x, val in enumerate(scanned_data[i]):
                print(val, x, n)
                output_sheet.cell(x + 2, n).value = val
        except Exception as e:
            print(e)
        n += 3
    c = True
    print(rf'{current.get()} return data.xlsx', os.listdir(r'C:\Users\Administrator\OneDrive\DAILY PHYSICAL RETURNS DATA\manual'))
    if rf'{current.get()} return data.xlsx' in os.listdir(r'C:\Users\Administrator\OneDrive\DAILY PHYSICAL RETURNS DATA\manual'):
        c = askyesno('File already exists', 'Do you want to replace old file with new one?')
    if c:
        output_file.save(rf'C:\Users\Administrator\OneDrive\DAILY PHYSICAL RETURNS DATA\manual\{current.get()} return data.xlsx')
        # ecom_yellow_chit_fixer.process(rf'C:\Users\neetu\OneDrive\DAILY PHYSICAL RETURNS DATA\manual\{current.get()} return data.xlsx')
        end_text['text'] = 'Done!!'
        end_text['fg'] = 'green'


root = Tk()

text_entry = Entry(root, width=30)
text_entry.pack()
today = datetime.date.today()
yesterday = today-datetime.timedelta(days=1)
day_before_yesterday = today-datetime.timedelta(days=2)

dates = [yesterday.__str__() + '    ' + yesterday.strftime("%A"), day_before_yesterday.__str__() + '    ' + day_before_yesterday.strftime("%A")]

current = StringVar(root)
current.set(dates[0])
dropdown = OptionMenu(root, current, *dates)
dropdown.pack()

listbox = Listbox(root, height=12, width=50, bg='black', fg='white', font=5)
listbox.pack()

finishButton = Button(root, text='Finish', command=finish)
finishButton.pack()

end_text = Label(root)
end_text.pack()

df = pd.read_excel('courier_classify.xlsx', sheet_name='Sheet1')
df_start = {}
df_end = {}
output_file = Workbook()
output_sheet = output_file.worksheets[0]

for x in df.index:
    for i in str(df.iloc[x, 1]).split(';'):
        if i != 'nan':
            df_start[i] = df.iloc[x, 0]

for x in df.index:
    for i in str(df.iloc[x, 2]).split(';'):
        if i != 'nan':
            df_end[i] = df.iloc[x, 0]
print(df_start, df_end)
scanned_data = {x: [] for x in df.iloc[:, 0]}
print(scanned_data, 'kjbkjbjkhb')
c = 0
update()

while True:
    text = text_entry.get()
    if text != '':
        c += 1
        if c == 7000:
            c = 0
            if len(text) > 5:
                a = ['codeg']
                for number in df_start:
                    if text.startswith(number):
                        a.append(df_start[number])
                for number in df_end:
                    if text.endswith(number):
                        a.append(df_end[number])
                if a[-1] != 'codeg':
                    pass_sound.play()
                    scanned_data[a[-1]].append(text)
                update()
                text_entry.delete(0, END)
                print(scanned_data)
            else:
                error_sound.play()
                text_entry.delete(0, END)
    root.update()